#!/usr/bin/env python3
"""
Extract ALL hyperlinks from the Roadmap XLSM spreadsheet.
Produces resource-index.json with every document link mapped to its task/sector.

Usage:
    python3 scripts/extract-links-from-xlsm.py /path/to/Roadmap.xlsm
"""

import sys
import json
import re
import os
from collections import OrderedDict

import openpyxl

# Sector sheets and their formats
SECTOR_SHEETS = {
    'Response Management': {'id': 'rmie', 'format': 'standard'},
    'Response Management (WIP)': {'id': 'rmie', 'format': 'standard'},
    'RMiE': {'id': 'rmie', 'format': 'rmie'},
    'Safeguarding': {'id': 'safeguarding', 'format': 'standard'},
    'Finance': {'id': 'finance', 'format': 'standard'},
    'Integra Launch': {'id': 'integra', 'format': 'integra'},
    'People & Culture': {'id': 'people_culture', 'format': 'standard'},
    'PCiE': {'id': 'people_culture', 'format': 'pcie'},
    'Supply Chain': {'id': 'supply_chain', 'format': 'standard'},
    'Safety & Security': {'id': 'safety_security', 'format': 'standard'},
    'Technical Programs': {'id': 'technical_programs', 'format': 'standard'},
    'MEAL': {'id': 'meal', 'format': 'standard'},
    'Grants': {'id': 'grants', 'format': 'standard'},
    'Partnerships': {'id': 'partnerships', 'format': 'standard'},
}

SECTOR_NAMES = {
    'rmie': 'Response Management',
    'finance': 'Finance',
    'people_culture': 'People & Culture',
    'supply_chain': 'Supply Chain',
    'safety_security': 'Safety & Security',
    'safeguarding': 'Safeguarding',
    'technical_programs': 'Technical Programs',
    'meal': 'MEAL',
    'grants': 'Grants',
    'partnerships': 'Partnerships',
    'integra': 'Integra Launch',
}


def is_valid_url(url):
    """Check if a URL is a real external link (not a relative ref or placeholder)."""
    if not url:
        return False
    url = url.strip()
    return url.startswith('http://') or url.startswith('https://') or url.startswith('mailto:')


def clean_name(name):
    """Clean up resource name from cell value."""
    if not name:
        return ''
    name = str(name).strip()
    # Remove leading/trailing whitespace and newlines
    name = re.sub(r'\s+', ' ', name).strip()
    return name


def find_header_row(ws, fmt):
    """Find the header row index for a sheet."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
        text = ' '.join(str(c.value or '') for c in row).lower()
        if fmt == 'rmie' or fmt == 'pcie':
            if 'task' in text and ('id' in text or 'title' in text):
                return i
        elif fmt == 'integra':
            if 'status' in text and 'task' in text:
                return i
        else:
            if 'response stage' in text and ('task' in text or 'subtask' in text or 'classification' in text):
                return i
            # Alternate: some standard sheets use different header patterns
            if 'tasks' in text and 'subtasks' in text and 'resources' in text:
                return i
    return None


def extract_sector_links(ws, sheet_name, config):
    """Extract all resource links from a sector sheet."""
    fmt = config['format']
    sector_id = config['id']
    resources = []

    header_row = find_header_row(ws, fmt)
    if not header_row:
        print(f'  Warning: No header found in {sheet_name}')
        return resources

    # Determine which columns contain resources/links based on format
    # We scan ALL cells for hyperlinks, associating them with the nearest task

    current_task = ''

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
        # Identify the task title from the appropriate column
        if fmt == 'rmie' or fmt == 'pcie':
            # RMiE/PCiE: col B=task ID, col C=task title, col D=resource link
            task_title = clean_name(row[2].value) if len(row) > 2 else ''  # col C (0-indexed: 2)
            if task_title and not task_title.startswith('['):
                current_task = task_title
            # Check col D for hyperlinks (resource link column)
            if len(row) > 3:
                cell = row[3]  # col D
                link_url = cell.hyperlink.target if cell.hyperlink else None
                link_name = clean_name(cell.value)
                if link_name and is_valid_url(link_url):
                    resources.append({
                        'name': link_name,
                        'url': link_url.strip(),
                        'sector': SECTOR_NAMES[sector_id],
                        'task': current_task or sheet_name,
                    })
        elif fmt == 'integra':
            # Integra: col D=task, col E=subtask, col H=resource
            task_title = clean_name(row[3].value) if len(row) > 3 else ''  # col D
            if task_title and not task_title.startswith('['):
                current_task = task_title
            # Check cols D, E, H for hyperlinks
            for ci in [3, 4, 7]:
                if len(row) > ci:
                    cell = row[ci]
                    link_url = cell.hyperlink.target if cell.hyperlink else None
                    link_name = clean_name(cell.value)
                    if link_name and is_valid_url(link_url):
                        resources.append({
                            'name': link_name,
                            'url': link_url.strip(),
                            'sector': SECTOR_NAMES[sector_id],
                            'task': current_task or sheet_name,
                        })
        else:
            # Standard format: col I=task, col J=subtask, col K=resource, col L=box link
            task_title = ''
            # Find task column (usually I, index 8)
            if len(row) > 8:
                task_title = clean_name(row[8].value)
            if task_title and not task_title.startswith('['):
                current_task = task_title

            # Check K (resource name) and L (box link) columns for hyperlinks
            # Also check all columns — some sheets have links in unexpected places
            for ci in range(len(row)):
                cell = row[ci]
                if not cell.hyperlink:
                    continue
                link_url = cell.hyperlink.target
                if not is_valid_url(link_url):
                    continue
                link_name = clean_name(cell.value)
                if not link_name:
                    continue
                # Skip contact emails in the resource index
                if link_url.startswith('mailto:'):
                    continue
                # Skip example/placeholder rows
                if any(x in link_name.lower() for x in ['[task', '[subtask', '[link', 'example', '[template']):
                    continue
                resources.append({
                    'name': link_name,
                    'url': link_url.strip(),
                    'sector': SECTOR_NAMES[sector_id],
                    'task': current_task or sheet_name,
                })

    return resources


def extract_service_links(ws):
    """Extract links from EmU Services sheet."""
    resources = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.hyperlink and cell.hyperlink.target:
                url = cell.hyperlink.target.strip()
                name = clean_name(cell.value)
                if name and is_valid_url(url) and not url.startswith('mailto:'):
                    resources.append({
                        'name': name,
                        'url': url,
                        'sector': 'Emergency Unit',
                        'task': 'EmU Services',
                    })
    return resources


def extract_preparedness_links(ws):
    """Extract links from Preparedness Library sheet."""
    resources = []
    current_category = ''
    for row in ws.iter_rows(min_row=2, values_only=False):
        # Col B = category, Col C = link
        cat_val = clean_name(row[1].value) if len(row) > 1 else ''
        link_cell = row[2] if len(row) > 2 else None

        if cat_val and not (link_cell and link_cell.value):
            current_category = cat_val

        if link_cell and link_cell.hyperlink and link_cell.hyperlink.target:
            url = link_cell.hyperlink.target.strip()
            name = clean_name(link_cell.value)
            if name and is_valid_url(url):
                resources.append({
                    'name': name,
                    'url': url,
                    'sector': 'Preparedness',
                    'task': current_category or 'Preparedness Library',
                })
    return resources


def main():
    if len(sys.argv) < 2:
        print('Usage: python3 extract-links-from-xlsm.py /path/to/Roadmap.xlsm')
        sys.exit(1)

    xlsm_path = sys.argv[1]
    print(f'Reading {xlsm_path}...')
    wb = openpyxl.load_workbook(xlsm_path, data_only=False)

    all_resources = []

    # Extract from sector sheets
    for sheet_name, config in SECTOR_SHEETS.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            links = extract_sector_links(ws, sheet_name, config)
            print(f'  {sheet_name}: {len(links)} resource links')
            all_resources.extend(links)

    # Extract from EmU Services
    if 'EmU Services' in wb.sheetnames:
        links = extract_service_links(wb['EmU Services'])
        print(f'  EmU Services: {len(links)} resource links')
        all_resources.extend(links)

    # Extract from Preparedness Library
    if 'Preparedness Library' in wb.sheetnames:
        links = extract_preparedness_links(wb['Preparedness Library'])
        print(f'  Preparedness Library: {len(links)} resource links')
        all_resources.extend(links)

    # Deduplicate by URL (keep first occurrence)
    seen_urls = set()
    unique_resources = []
    for r in all_resources:
        url_key = r['url'].lower().rstrip('/')
        if url_key not in seen_urls:
            seen_urls.add(url_key)
            unique_resources.append(r)

    print(f'\nTotal: {len(all_resources)} links found, {len(unique_resources)} unique')

    # Write resource-index.json
    out_dir = os.path.join(os.path.dirname(__file__), '..', 'src', 'data')
    out_path = os.path.join(out_dir, 'resource-index.json')
    with open(out_path, 'w') as f:
        json.dump(unique_resources, f, indent=2)
    print(f'Wrote {out_path} ({len(unique_resources)} resources)')

    # Also copy to public/
    public_path = os.path.join(os.path.dirname(__file__), '..', 'public', 'resource-index.json')
    with open(public_path, 'w') as f:
        json.dump(unique_resources, f, indent=2)
    print(f'Wrote {public_path}')


if __name__ == '__main__':
    main()
